from flask import Flask, request, render_template, Response
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows
import io


app = Flask(__name__)


@app.route('/', methods=['GET', 'POST'])
def convert():
    if request.method == 'GET':
        return render_template('index.html')

    if request.method == 'POST':
        file = request.files['file']

        # Загрузка файла с парсингом колонок 'Филиал', 'Сотрудник', 'Налоговая база', 'Исчислено всего'
        # и исключением первой и последней строки
        df = pd.read_excel(
            io=file,
            header=1,
            usecols="A, B, E, F"
        )[:-1]

        # Удаление пустой строки, сброс индексов, переименование
        df = df.drop(labels=[0], axis=0)
        df = df.reset_index(drop=True)
        df.columns = ['Филиал', 'Сотрудник', 'Налоговая база', 'Исчислено всего',]

        # Расчет колонки 'Исчислено всего по формуле' с условием
        df['Исчислено всего по формуле'] = np.where(
            df['Налоговая база'] < 5000000,
            df['Налоговая база'] * 13/100,
            df['Налоговая база'] * 15/100
        )

        # Расчет колонки 'Отклонения'
        df['Отклонения'] = df['Исчислено всего'] - df['Исчислено всего по формуле']

        # Сортировка 'Отклонения' по убыванию
        df = df.sort_values(by='Отклонения', ascending=False)

        # Создание Workbook из DataFrame
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Форматирование таблицы
        ws.insert_rows(0, 1)
        title = [
            {'col': 'A', 'width': 40.22},
            {'col': 'B', 'width': 34.22},
            {'col': 'C', 'width': 13.33},
            {'col': 'D', 'width': 13.33},
            {'col': 'E', 'width': 13.33},
            {'col': 'F', 'width': 16.44},
        ]

        def stylizationCell(col, str):
            ws[f'{col}{str}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f'{col}{str}'].fill = PatternFill('solid', fgColor="CBE4E5")
            ws[f'{col}{str}'].font = Font(bold=True, color='000000', name='Arial', size=10)
            side = Side(border_style="thin", color="000000")
            ws[f'{col}{str}'].border = Border(top=side, bottom=side, left=side, right=side)

        for el in title:
            value = ws[f'{el["col"]}2'].value
            stylizationCell(el["col"], 1)
            ws.merge_cells(f'{el["col"]}1:{el["col"]}2')
            ws[f'{el["col"]}1'].value = value
            ws.column_dimensions[el["col"]].width = el["width"]

        ws.unmerge_cells('D1:D2')
        ws.unmerge_cells('E1:E2')
        ws['D2'].value = ws['D1'].value
        ws['E2'].value = ws['E1'].value
        stylizationCell('E', 2)
        stylizationCell('D', 2)
        ws.row_dimensions[2].height = 40

        ws.merge_cells('D1:E1')
        ws['D1'].value = 'Налог'
        stylizationCell('D', 1)

        # Индикация ячеек с отклонением
        for cell in ws['F'][2:]:
            try:
                if cell.value == 0:
                    cell.fill = PatternFill('solid', fgColor="90EE90")
                else:
                    cell.fill = PatternFill('solid', fgColor="FF6347")
            except ValueError:
                print('Err')

        # wb.save("static/Отчет.xlsx")

        buffer = io.BytesIO()
        wb.save(buffer)

        headers = {
            'Content-Disposition': 'attachment; filename=output.xlsx',
            'Content-type': 'application/vnd.ms-excel'
        }

        return Response(buffer.getvalue(), mimetype='application/vnd.ms-excel', headers=headers)






if __name__ == '__main__':
    app.run()
